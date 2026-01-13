#insatll openpyxl usinf pip install 
# pip install openpyxl

#WAP to reaad a cell from excel file.
import openpyxl
def read_file_excel(File_path,Sheet_name,Cell_name):
    work_Book=openpyxl.load_workbook(File_path)
    sheet=work_Book[Sheet_name]
    cell_value=sheet[Cell_name].value
    print(f"The value of cell{Cell_name} is :{cell_value}")

#read_file_excel(File_path="User_File.xlsx",Sheet_name="Sheet1",Cell_name="A7")

#Writea data into excel sheet
def write_file_excel(file_path,Sheet_name,cell_name,data):
    workbook=openpyxl.load_workbook(file_path)
    #open sheet
    sheet=workbook[Sheet_name]
    #write dat to cell
    sheet[cell_name]=data
    print(f"Write a file to {file_path} in {Sheet_name} at {cell_name} with data:{data}")
    workbook.save(file_path)


#New_data="Hello World"
#write_file_excel(file_path="User_file.xlsx",Sheet_name="Sheet1",cell_name="C4",data=New_data)
'''
data_list = ['Data1', 'Data2', 'Data3', 'Data4', 'Data5']
for index, val in enumerate(data_list):
    print("Index is :",index,"Value is :",val)
    write_file_excel(file_path='User_File.xlsx', Sheet_name='Sheet3', cell_name=f'E{index+1}', 
                     data=val)
'''

'''
credentials_list = [
    ['Username', 'Password'],  
    ['user1', 'pass1'],
    ['user2', 'pass2'],
    ['user3', 'pass3'],
    ['user4', 'pass4'],
]

for index,val in enumerate(credentials_list):
    write_file_excel(file_path='User_File.xlsx', Sheet_name='Sheet3', cell_name=f'A{index+1}',
                     data=val[0])
    write_file_excel(file_path='User_File.xlsx', Sheet_name='Sheet3', cell_name=f'B{index+1}',
                     data=val[1])

'''
Student_details=[
    ['NAME','AGE','GRADE'],
    ['RAM','13','A'],
    ['SITA','14','B'],
    ['GITA','13','A+'],
    ['JOHN','15','C']]
'''
#print(Sudent_details)
for index,val in enumerate(Student_details):
    write_file_excel(file_path='User_File.xlsx',Sheet_name='Sheet4',cell_name=f'A{index+1}',
                     data=val[0])
    write_file_excel(file_path='User_File.xlsx',Sheet_name='Sheet4',cell_name=f'B{index+1}',
                     data=val[1])
    write_file_excel(file_path='User_File.xlsx',Sheet_name='Sheet4',cell_name=f'C{index+1}',
                     data=val[2])   

'''
########################
# get all rows and colums from excel sheet
def read_all_data_excel(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columns
    #for row in sheet.iter_rows(values_only=True):
    #   print(row)
    for column in sheet.iter_cols(values_only=True):
       print(column)

#read_all_data_excel(file_path='User_File.xlsx', sheet_name='Sheet2')

def read_all_data_excel1(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columnsgit 
    max_row = sheet.max_row # maximum number of rows
    print(max_row)
    max_column = sheet.max_column # maximum number of columns
    print(max_column)
    for i in range(2, max_row + 1): #i is about rows
        for j in range(1, max_column + 1): # j is about columns
            cell_value = sheet.cell(row=i, column=j).value
            print(cell_value, end=" | ")
        print()  # New line after each row
        

read_all_data_excel1(file_path='User_File.xlsx', sheet_name='Sheet2')