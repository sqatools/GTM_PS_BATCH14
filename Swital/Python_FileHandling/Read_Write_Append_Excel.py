# install openpyxl package if not already installed
# pip install openpyxl
import openpyxl
def read_excel_file(file_path, sheet_name, cell_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Read the value from the specified cell
    cell_value = sheet[cell_name].value
    print(f"Value in {cell_name} of sheet '{sheet_name}': {cell_value}")
# read one cell data 
read_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet1', cell_name='A2')

# get all rows and columns from excel sheet
#for i in range(1, 6):
#    read_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet1', cell_name=f'A{i}')

#### Write data to excel file
def write_excel_file(file_path, sheet_name, cell_name, value):  
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Write the value to the specified cell
    sheet[cell_name] = value
    # Save the workbook
    workbook.save(file_path)
    print(f"Wrote '{value}' to {cell_name} of sheet '{sheet_name}'")
# write one cell data
#write_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet1', cell_name='B2', value='Hello, World!')

# Write multiple cell data
#data_list = ['Data1', 'Data2', 'Data3', 'Data4', 'Data5']
#for i in range(1, 6):
#    write_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet2', cell_name=f'D{i}', value=data_list[i-1])

# Write multiple rows and columns
credentials_list = [
    ['User1', 'Pass1'],
    ['User2', 'Pass2'],
    ['User3', 'Pass3'],
    ['User4', 'Pass4'],
    ['User5', 'Pass5']
]
#for index, val in enumerate(credentials_list):
#    write_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet3', cell_name=f'A{index+1}', value=val[0])
#    write_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet3', cell_name=f'B{index+1}', value=val[1])

# get all rows and colums from excel sheet
def read_all_data_excel(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Iterate through all rows and columns
    for row in sheet.iter_rows(values_only=True):
        print(row)
#read_all_data_excel(file_path='Read_Excel.xlsx', sheet_name='Sheet2')

# Read cell data using for loop
for cell in "ABCDE":
    read_excel_file(file_path='Read_Excel.xlsx', sheet_name='Sheet3', cell_name=f'{cell}2')

# Read all data from excel sheet upto max rows and columns
def read_all_data_excel(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]

    
    # Iterate through all rows and columns
    max_row = sheet.max_row
    max_column = sheet.max_column
    for i in range(2, max_row + 1):
        for j in range(1, max_column + 1):
            cell_value = sheet.cell(row=i, column=j).value
            print(cell_value, end=" | ")
        print()  # New line after each row
#read_all_data_excel(file_path='Read_Excel.xlsx', sheet_name='Sheet3')

# using iter_rows
def read_all_data_excel(file_path, sheet_name): 
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columns
    rows_values = list(sheet.iter_rows(values_only=True))
    for row in rows_values[1:]:  # Skip header row
        print(row)
#read_all_data_excel(file_path='Read_Excel.xlsx', sheet_name='Sheet1')