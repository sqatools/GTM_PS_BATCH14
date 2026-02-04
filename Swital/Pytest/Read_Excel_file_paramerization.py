import openpyxl

def get_test_data(file_path, sheet_name):

    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]

    data = []

    # Start from row 2 (assuming row 1 = header)
    for i in range(2, sheet.max_row + 1):

        username = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value

        data.append((username, password))  # tuple

    wb.close()

    return data
