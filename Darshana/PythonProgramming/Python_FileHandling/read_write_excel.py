# install openpyxl package if not already installed
# pip install openpyxl
import openpyxl

def read_excel_file(file_path, sheet_name, cell_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Read the value from the specified cell
    cell_value = sheet[cell_name].value
    print(f"Value in {cell_name} of sheet '{sheet_name}': {cell_value}")

# read one cell data 
#read_excel_file(file_path='users_data.xlsx', sheet_name='Sheet1', cell_name='A2')

# read multiple cell data
# for i in range(1, 6):
#     read_excel_file(file_path='users_data.xlsx', sheet_name='Sheet1', cell_name=f'A{i}')


def write_excel_file(file_path, sheet_name, cell_name, data):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Write data to the specified cell
    sheet[cell_name] = data
    
    # Save the workbook
    workbook.save(file_path)
    print(f"Data written to sheet '{sheet_name}' starting at {cell_name}.")


"""
write_excel_file(file_path='users_data.xlsx', sheet_name='Sheet1', cell_name='D2', data='Learning Excel with Python')

data_list = ['Data1', 'Data2', 'Data3', 'Data4', 'Data5']
for index, val in enumerate(data_list):
    write_excel_file(file_path='users_data.xlsx', sheet_name='Sheet3', cell_name=f'E{index+1}', data=val)

credentials_list = [
    ['Username', 'Password'],  
    ['user1', 'pass1'],
    ['user2', 'pass2'],
    ['user3', 'pass3'],
    ['user4', 'pass4'],
]

for index, val in enumerate(credentials_list):
    write_excel_file(file_path='users_data.xlsx', sheet_name='Sheet2', cell_name=f'A{index+1}', data=val[0])
    write_excel_file(file_path='users_data.xlsx', sheet_name='Sheet2', cell_name=f'B{index+1}', data=val[1])
"""

# get all rows and colums from excel sheet
def read_all_data_excel(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columns
    for row in sheet.iter_rows(values_only=True)[0:]:
        print(row)

# print("#"*50)
# #read_all_data_excel(file_path='users_data.xlsx', sheet_name='Sheet4')
# for cell in "ABCDEFGHIJ":
#     read_excel_file(file_path='users_data.xlsx', sheet_name='Sheet1', cell_name=f'{cell}1')

#read_all_data_excel(file_path='users_data.xlsx', sheet_name='Sheet2')


def read_all_data_excel(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columns
    max_row = sheet.max_row # maximum number of rows
    max_column = sheet.max_column # maximum number of columns
    for i in range(2, max_row + 1):
        for j in range(1, max_column + 1):
            cell_value = sheet.cell(row=i, column=j).value
            print(cell_value, end=" | ")
        print()  # New line after each row
        

#read_all_data_excel(file_path='users_data.xlsx', sheet_name='Sheet2')


def read_all_data_excel_new(file_path, sheet_name):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    
    # Iterate through all rows and columns
    rows_values = list(sheet.iter_rows(values_only=True))
    # print(rows_values)
    for row in rows_values[1:]:
        print(row)

#read_all_data_excel_new(file_path='users_data.xlsx', sheet_name='Sheet2')



def write_excel_file_with_multiple_values(file_path, sheet_name, cell_list, data_list):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    # Select the specified sheet
    sheet = workbook[sheet_name]
    # Write data to the specified cell
    for i in range(len(cell_list)):
        sheet[cell_list[i]] = data_list[i]
    
    # Save the workbook
    workbook.save(file_path)


cell_list = ['A1', 'B1', 'C1', 'D1', 'A2', 'B2', 'C2']
data_list1 = [20, 30, 40, 50, 60, 70, 80]

write_excel_file_with_multiple_values(file_path="users_data.xlsx", sheet_name="Sheet5", cell_list=cell_list, data_list=data_list1)