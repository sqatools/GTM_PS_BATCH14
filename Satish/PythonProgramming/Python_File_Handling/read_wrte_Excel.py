# install openpyxl package if not already installed
# pip install openpyxl
import openpyxl

def read_excel_file(file_path,sheet_name):
    # Load the workbook
    workbook=openpyxl.load_workbook(file_path,data_only=True) # Load workbook with computed  values instead of formula then use data_only=True
    #Select the specified sheet
    sheet=workbook[sheet_name] # or we can use sheet=workbook.active

    for row in sheet.iter_rows(min_row=2,min_col=1,max_col=7,values_only=True): # To get value of cell we use  values_only=True. iter_rows function used to read rows only not for write.
        print(row)

read_excel_file('test.xlsx','Sheet1')


def write_excel_file(file_path,sheet_name,cell_list,data_list):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    for index,cell in enumerate(cell_list):
        sheet[cell]=data_list[index]
        workbook.save(file_path)

cell_list1=['K1','L1','M1','N1']
data_list1=[10,20,20,40]

write_excel_file('test.xlsx','Sheet1',cell_list1, data_list1)
write_excel_file('test2.xlsx', 'Sheet2', ['A1'], [30] )
